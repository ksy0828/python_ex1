from flask import Flask, render_template, request, send_file
import os
from openpyxl import load_workbook
from deep_translator import GoogleTranslator

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['GET', 'POST'])
def upload():
    file = request.files["file"] #파일은 리퀘스트로 받음
    file.save(os.path.join("uploads", file.filename))

    #엑셀 파일 읽기
    wb = load_workbook(os.path.join("uploads", file.filename))
    ws = wb.active

    for row in ws.iter_rows():
        for cell in row:
            translated_text = GoogleTranslator(source='ko', target='en').translate(cell.value)
            cell.value = translated_text

    wb.save('result_en.xlsx')        
    return render_template('result.html', filename=file.filename)

@app.route('/download_report')
def download_report():
    return send_file('result_en.xlsx', as_attachment=True)
    
if __name__ == '__main__':
    app.run(debug=True)
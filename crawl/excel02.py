# 앞에서 진행했던 malwares 트래픽 사이트의 결과를 엑셀 파일로 저장하시오.
# 1. "설명", "링크"를 헤더로 미리 입력
# 2. 그 아래로 데이터들을 저장

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

url = "https://www.malware-traffic-analysis.net/2023/index.html"
headers = {'User-Agent': 'Mozilla/5.0',
        'Content-Type': 'text/html; charset=utf-8'}
req = requests.get(url, headers=headers)
soup = BeautifulSoup(req.text, "lxml")
tags = soup.select('#main_content > div.content > ul > li > a.main_menu')
#print(tags)

wb = Workbook()
ws = wb.active #현재 활성화된 시트에 접근 (마지막으로 편집한 곳)
ws['A1'] = "설명"
ws['B1'] = "URL 링크"

# 링크 정보 추출 및 출력
i=2
for tag in tags:
    ws.cell(row=(i), column=1, value=tag.text)
    ws.cell(row=(i), column=2, value=f"https://www.malware-traffic-analysis.net/2023/{tag['href']}")
    i = i+1

wb.save("malwaresQ.xlsx")
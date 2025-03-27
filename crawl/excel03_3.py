from openpyxl import load_workbook

# 엑셀 파일 불러오기
wb = load_workbook("excel_data2.xlsx", data_only=True) #date_only:데이터만 받겠다(수식x)
ws = wb.active

cell = ws["A1":"E7"]

for row in ws["A1":"E7"]:
    result = []
    for cell in row:
        result.append(cell.value)
    print(result)

for row in ws["A1":"E7"]:
    # 리스트 컴프리헨션으로 결과를 리스트로 저장
    values = [cell.value for cell in row]
    print(values)
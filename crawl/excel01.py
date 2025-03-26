from openpyxl import Workbook
wb = Workbook()

# ws1 = wb.create_sheet("Mysheet")
ws = wb.active
ws.title = 'New Titles' #워크시트 제목

for row in ws['A1:D4']:
    print(f"row:{row}")
    for cell in row:
        print(f"cell:{cell}")
        cell.value = '헬로우 파이썬'

wb.save('balance.xlsx') #파일 제목.확장자


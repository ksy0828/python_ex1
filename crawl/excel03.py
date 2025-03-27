from openpyxl import load_workbook

# 파일을 불러와서 워크북 객체 생성
wb = load_workbook("excel_data.xlsx")

# 워크시트 이름으로 접근
ws = wb['Sheet1']  # 'Sheet1'이라는 이름의 워크시트에 접근

# D18 셀의 값을 읽음
cell_value = ws['A1'].value
print(f"value : {cell_value}")
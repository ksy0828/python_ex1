from openpyxl import load_workbook

# 엑셀 파일 불러오기
wb = load_workbook("excel_data.xlsx")
ws = wb.active

cell = ws["A1":"C7"]

for row in ws["A1":"C7"]:
    result = []
    for cell in row:
        result.append(cell.value)
    print(result)

#line10~12를 line18로 한줄로 작성 근데 익숙하지 않으면 안쓰는게
for row in ws["A1":"C7"]:
    # 리스트 컴프리헨션으로 결과를 리스트로 저장
    values = [cell.value for cell in row]
    print(values)